import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import Image
from PIL import Image, ImageTk
from openpyxl import load_workbook
import tkinter.font as tkFont
import math
import openpyxl
import csv
import json
import xlwt



#THIS SOFTWARE BELONGS TO AND IS OWNED BY CHIEF ENGINEER OLOROGUN ENOCH O. EJOFODOMI AND ENGINEER  FRANCIS OLAWUYI.
#THIS SOFTWARE IS OWNED BY DR. JASON ZARA, DR. VESNA ZDERIC, ENGINEER FRANCIS OLAWUYI, DR MICHAEL OLAWUYI, DR. MATTHEW OLAWUYI, DR. ESTHER OLAWUYI, DR. AHMED JENDOUBI, DR. MOHAMED  CHOUIKHA, ENGINEER  ENOCH  EJOFODOMI, EFEJERA EJOFODOMI, AND LUCKY EJOFODOMI.
#THIS SOFTWARE IS A MEDICAL SOFTWARE THAT ALLOWS PHYSICIANS  TO KEEP  ELECTRONIC  RECORDS OF PATIENTS.
#November  13, 2023.

#create main window of the application
mainwindow = tk.Tk()
mainwindow.geometry("1400x700")
mainwindow.title('AD CAD SOFTWARE BY OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218')
mainwindow['background']='#7F7FFF'
mainwindow.iconbitmap('companylogo.ico')

global textBox1
global textBox2
global row


def AddPData2():

    global textBox1a
    global textBox2a
    global textBox3a
    global textBox4a
    global textBox5a
    global textBox6a
    global textBox7a
    global textBox8a
    global textBox9a
    global foundpatientrow2
    global n

    
    data1=(textBox1a.get("1.0","end-1c")) #print(length)
    data2=(textBox2a.get("1.0","end-1c"))  #print(width)
    data3=(textBox3a.get("1.0","end-1c")) #print(length)
    data4=(textBox4a.get("1.0","end-1c"))  #print(width)
    data5=(textBox5a.get("1.0","end-1c")) #print(length)
    data6=(textBox6a.get("1.0","end-1c"))  #print(width)
    data7=(textBox7a.get("1.0","end-1c")) #print(length)
    data8=(textBox8a.get("1.0","end-1c"))  #print(width)
    data9=(textBox9a.get("1.0","end-1c")) #print(length)

    if(data1 == ''):
       data1=int(0)
    if(data2 == ''):
       data2=int(0)       
    if(data3 == ''):
       data3=0
    if(data4 == ''):
       data4=0
    if(data5 == ''):
       data5=0       
    if(data6 == ''):
       data6=0
    if(data7 == ''):
       data7=0
    if(data8 == ''):
       data8=0
    if(data9 == ''):
       data9=0

    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("PATIENTDATA2.xlsx")
    # file = open('PATIENTDATA2.xlsx', 'r') 
    # Define variable to read sheet
    dataframe1 = dataframe.active
    dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    dataframe1.insert_rows(idx=foundpatientrow2 + n + 1)#- insert at 3rd position
    #dataframe1.append(dataB)
    dataframe.save("PATIENTDATA2.xlsx")
    aaa=1;
    for row in range(foundpatientrow2 + n + 1, foundpatientrow2 + n + 2):
        dataframe1.cell(row,1).value=data1
        dataframe1.cell(row,2).value=data2
        dataframe1.cell(row,3).value=data3
        dataframe1.cell(row,4).value=data4
        dataframe1.cell(row,5).value=data5
        dataframe1.cell(row,6).value=data6
        dataframe1.cell(row,7).value=data7
        dataframe1.cell(row,8).value=data8
        dataframe1.cell(row,9).value=data8
            
           
    dataframe.save("PATIENTDATA2.xlsx")                        
    # dataframe1.cell(row=dataframe1.max_row+1,column =1).value="TESTING "
    # dataframe1.getCells().insertColumns(1,1) - insert at 2nd position
    # dataframe1.getCells().insertRow(2,1) - insert at 3rd position
    # dataframe1.save("PATIENTDATA2.xlsx")
    
   # dataframe1.append([data1, data2, data3, data4, data5, data6, data7, data8, data9])
    #file = open('PATIENTDATA2.xlsx', 'w')
    #file = csv.writer(file)
    #file.writerow(['Name', 'Age', 'Enrollment Number'])
 
      #  dataframe1write_column(dataframe1.max_row+1, 1,data1)
    
    #ws = xw.Book("PATIENTDATA2.xlsx").sheets['Sheet1']
    #v1 = ws.range("C1:C7").value
    # v2 = ws.range("F5").value
   # print("Result:", v1)
   # ws.append([data1, data2, data3, data4, data5, data6, data7, data8, data9])

       #file = open('PATIEN,TDATA2.xlsx', 'w')
   # file.writerow([data1, data2, data3, data4, data5, data6, data7, data8, data9])
    print('All records inserted successfully !')
    print([data1])
    print([data2])
    print([data3])
    print([data4])
    print([data5])
    print([data6])
    print([data7])
    print([data8])
    print([data9])

    #create window
    tk3 = Toplevel()
    tk3.geometry("700x200")
    tk3.title('ELECTRONIC RECORD KEEPING BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(700,200)
    tk3.maxsize(700,200)

            
    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=50)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)

    #insert RC14668218 Company logo, El Elyon, UK FLAG and UK MAP
    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=4,y=1)
    img4b=ImageTk.PhotoImage(file="ELELYON.png")
    lb4b = Label(tk3, image=img4b)
    lb4b.img4b = img4b #Save reference to image
    lb4b.place(x=632,y=1)
    label4 = Label(tk3,text="PATIENT 'S DATA ENTERED  SUCESSFULLY", width=35, height=2, font=('Arial bold', 20)) 
    label4.place(x=50,y=120)



def AddPData():

    global textBox1a
    global textBox2a
    global textBox3a
    global textBox4a
    global textBox5a
    global textBox6a
    global textBox7a
    global textBox8a
    global textBox9a
    global foundpatientrow2
    global n

    
    data1=(textBox1a.get("1.0","end-1c")) #print(length)
    data2=(textBox2a.get("1.0","end-1c"))  #print(width)
    data3=(textBox3a.get("1.0","end-1c")) #print(length)
    data4=(textBox4a.get("1.0","end-1c"))  #print(width)
    data5=(textBox5a.get("1.0","end-1c")) #print(length)
    data6=(textBox6a.get("1.0","end-1c"))  #print(width)
    data7=(textBox7a.get("1.0","end-1c")) #print(length)
    data8=(textBox8a.get("1.0","end-1c"))  #print(width)
    data9=(textBox9a.get("1.0","end-1c")) #print(length)

    if(data1 == ''):
       data1=int(0)
    if(data2 == ''):
       data2=int(0)       
    if(data3 == ''):
       data3=0
    if(data4 == ''):
       data4=0
    if(data5 == ''):
       data5=0       
    if(data6 == ''):
       data6=0
    if(data7 == ''):
       data7=0
    if(data8 == ''):
       data8=0
    if(data9 == ''):
       data9=0

    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("PATIENTDATA2.xlsx")
    # file = open('PATIENTDATA2.xlsx', 'r') 
    # Define variable to read sheet
    dataframe1 = dataframe.active
    
    dataB = (data1, data2, data3, data4, data5, data6, data7, data8, data9)
    dataframe1.append(dataB)
    dataframe.save("PATIENTDATA2.xlsx")
    
    # dataframe1.cell(row=dataframe1.max_row+1,column =1).value="TESTING "
    # dataframe1.getCells().insertColumns(1,1) - insert at 2nd position
    # dataframe1.getCells().insertRow(2,1) - insert at 3rd position
    # dataframe1.save("PATIENTDATA2.xlsx")
    
   # dataframe1.append([data1, data2, data3, data4, data5, data6, data7, data8, data9])
    #file = open('PATIENTDATA2.xlsx', 'w')
    #file = csv.writer(file)
    #file.writerow(['Name', 'Age', 'Enrollment Number'])
 
      #  dataframe1write_column(dataframe1.max_row+1, 1,data1)
    
    #ws = xw.Book("PATIENTDATA2.xlsx").sheets['Sheet1']
    #v1 = ws.range("C1:C7").value
    # v2 = ws.range("F5").value
   # print("Result:", v1)
   # ws.append([data1, data2, data3, data4, data5, data6, data7, data8, data9])

       #file = open('PATIEN,TDATA2.xlsx', 'w')
   # file.writerow([data1, data2, data3, data4, data5, data6, data7, data8, data9])
    print('All records inserted successfully !')
    print([data1])
    print([data2])
    print([data3])
    print([data4])
    print([data5])
    print([data6])
    print([data7])
    print([data8])
    print([data9])

    #create window
    tk3 = Toplevel()
    tk3.geometry("700x200")
    tk3.title('ELECTRONIC RECORD KEEPING BY RC14668218')
    tk3['background']='#36454F'
    tk3.minsize(700,200)
    tk3.maxsize(700,200)

            
    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=50)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)

    #insert RC14668218 Company logo, El Elyon, UK FLAG and UK MAP
    img4=PhotoImage(file="COMPANYLOGO.png")
    lb4 = Label(tk3, image=img4)
    lb4.img4 = img4 #Save reference to image
    lb4.place(x=4,y=1)
    img4b=ImageTk.PhotoImage(file="ELELYON.png")
    lb4b = Label(tk3, image=img4b)
    lb4b.img4b = img4b #Save reference to image
    lb4b.place(x=632,y=1)
    label4 = Label(tk3,text="PATIENT'S  DATA ENTERED  SUCESSFULLY", width=35, height=2, font=('Arial bold', 20)) 
    label4.place(x=50,y=120)


def PAddNew():

    global textBox1
    global textBox2
    global textBox1a
    global textBox2a
    global textBox3a
    global textBox4a
    global textBox5a
    global textBox6a
    global textBox7a
    global textBox8a
    global textBox9a
    global foundpatientrow2
    global n

    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("PATIENTDATA2.xlsx")
    # file = open('PATIENTDATA2.xlsx', 'r') 
    # Define variable to read sheet
    dataframe1 = dataframe.active


    #create window
    tk3 = Toplevel()
    tk3.geometry("1200x1200")
    tk3.title('ELECTRONIC RECORD KEEPING BY OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218')
    tk3['background']='#36454F'

            
    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    # Load the images of RC14668218 LOGO AND EL ELYON
   # img2=PhotoImage(file="LOGO3.png")
   # lb2 = Label(tk3, image=img2)
   # lb2.img2 = img2 #Save reference to image
   # lb2.place(x=50,y=50)
   # img3=ImageTk.PhotoImage(file="ELELYONC.png")
   # lb3 = Label(tk3, image=img3)
   # lb3.img3 = img3 #Save reference to image
   # lb3.place(x=1320,y=50)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=50)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)



    #Insert RC14668218 COMPANY HEADING

    #lb3b = Label(tk3, image=img3b)
   # lb3b.img3b = img3b #Save reference to image
   # lb3b.place(x=1227,y=1)

    label2 = Label(tk3, text="MEDICAL SOFTWARE FOR ELECTRONIC RECORD KEEEPING", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 

    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)


    #Create Canvas for RC14668218 Irrigation Layout
    canvas5=Canvas(tk3,bg='#FFFFFF',width=1300,height=550)
    canvas5.place(x=35,y=120)
    

    img3=ImageTk.PhotoImage(file="ELELYONC.png")
    lb3 = Label(canvas5, image=img3)
    lb3.img3 = img3 #Save reference to image
    lb3.place(x=20,y=1)

    
    labell1 = Label(canvas5, text="PATIENT  NAME: ", width=17, height=2, font=('Arial bold', 20)) 
    labell1.place(x=200,y=20)
    textBox1a=Text(canvas5, height=4, width=40)
    textBox1a.place(x=520,y=20)
    canvas5.create_rectangle(518, 18, 848, 90,outline= "black", width = 6)

    labell2 = Label(canvas5, text="GENDER: ", width=8, height=2, font=('Arial bold', 20)) 
    labell2.place(x=900,y=20)
    textBox2a=Text(canvas5, height=4, width=20)
    textBox2a.place(x=1060,y=20)
    canvas5.create_rectangle(1058, 18, 1228, 90,outline= "black", width = 6)
    
    labell3 = Label(canvas5, text="LOCATION: ", width=17, height=2, font=('Arial bold', 20)) 
    labell3.place(x=200,y=110)
    textBox3a=Text(canvas5, height=4, width=40)
    textBox3a.place(x=520,y=110)
    canvas5.create_rectangle(518, 108, 848, 180,outline= "black", width = 6)

    labell4 = Label(canvas5, text="AGE: ", width=8, height=2, font=('Arial bold', 20)) 
    labell4.place(x=900,y=110)
    textBox4a=Text(canvas5, height=4, width=20)
    textBox4a.place(x=1060,y=110)
    canvas5.create_rectangle(1058, 108, 1228, 180,outline= "black", width = 6)

    labell5 = Label(canvas5, text="DATE OF VISIT: ", width=17, height=2, font=('Arial bold', 20)) 
    labell5.place(x=200,y=200)
    textBox5a=Text(canvas5, height=4, width=12)
    textBox5a.place(x=520,y=200)
    canvas5.create_rectangle(518, 198, 623, 270,outline= "black", width = 6)

    labell6 = Label(canvas5, text="PRE-EXISTING CONDITION: ", width=22, height=2, font=('Arial bold', 20)) 
    labell6.place(x=640,y=200)
    textBox6a=Text(canvas5, height=4, width=23)
    textBox6a.place(x=1040,y=200)
    canvas5.create_rectangle(1038, 198, 1230, 270,outline= "black", width = 6)


    labell7 = Label(canvas5, text=" DISEASE(S) DIAGNOSED: ", width=20, height=2, font=('Arial bold', 20)) 
    labell7.place(x=5,y=290)
    textBox7a=Text(canvas5, height=4, width=32)
    textBox7a.place(x=360,y=290)
    canvas5.create_rectangle(358, 288, 623, 360,outline= "black", width = 6)

    labell8 = Label(canvas5, text="DRUG(S) PRESCRIBED: ", width=19, height=2, font=('Arial bold', 20)) 
    labell8.place(x=640,y=290)
    textBox8a=Text(canvas5, height=4, width=30)
    textBox8a.place(x=985,y=290)
    canvas5.create_rectangle(983, 288, 1230, 360,outline= "black", width = 6)

    labell9 = Label(canvas5, text="PREVIOUS HEALTH HISTORY: ", width=24, height=2, font=('Arial bold', 20)) 
    labell9.place(x=5,y=400)
    textBox9a=Text(canvas5, height=8, width=64)
    textBox9a.place(x=440,y=390)
    canvas5.create_rectangle(437, 388, 964, 530,outline= "black", width = 6)

    button3 = Button(canvas5, text='ADD DATA',font=('Arial bold', 36), width=10, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=AddPData2)
    button3.place(x=980,y=410)

 
   # labell3 = Label(canvas5, text="AGE ", width=5, height=2, font=('Arial bold', 20)) 
   # labell3.place(x=665,y=20)
   # labell4 = Label(canvas5, text="LOCATION ", width=14, height=2, font=('Arial bold', 20)) 
   # labell4.place(x=770,y=20)    
   # labell5 = Label(canvas5, text="DATE OF PHYSICIAN VISIT ", width=30, height=2, font=('Arial bold', 20)) 
   # labell5.place(x=975,y=20)    
   # labell6 = Label(canvas5, text="DISEASE(S) DIAGNOSED ", width=30, height=2, font=('Arial bold', 20)) 
   # labell6.place(x=1165,y=20)    
   # labell7 = Label(canvas5, text="DRUGS PRESCRIBED ", width=40, height=2, font=('Arial bold', 20)) 
   # labell7.place(x=1465,y=20)    
   # labell8 = Label(canvas5, text="PREVIOUS HEALTH HISTORY ", width=60, height=2, font=('Arial bold', 20)) 
  #  labell8.place(x=1765,y=20)    
   # labell9 = Label(canvas5, text="PRE-EXISTING MEDICAL  CONDITION(S)", width=40, height=2, font=('Arial bold', 20)) 
   # labell9.place(x=2065,y=20)


    

def PAdd():

    global textBox1
    global textBox2
    global textBox1a
    global textBox2a
    global textBox3a
    global textBox4a
    global textBox5a
    global textBox6a
    global textBox7a
    global textBox8a
    global textBox9a
    global foundpatientrow2
    global n

    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("PATIENTDATA2.xlsx")
    # file = open('PATIENTDATA2.xlsx', 'r') 
    # Define variable to read sheet
    dataframe1 = dataframe.active


    #create window
    tk3 = Toplevel()
    tk3.geometry("1200x1200")
    tk3.title('ELECTRONIC RECORD KEEPING BY OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218')
    tk3['background']='#36454F'

            
    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    # Load the images of RC14668218 LOGO AND EL ELYON
   # img2=PhotoImage(file="LOGO3.png")
   # lb2 = Label(tk3, image=img2)
   # lb2.img2 = img2 #Save reference to image
   # lb2.place(x=50,y=50)
   # img3=ImageTk.PhotoImage(file="ELELYONC.png")
   # lb3 = Label(tk3, image=img3)
   # lb3.img3 = img3 #Save reference to image
   # lb3.place(x=1320,y=50)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=50)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)



    #Insert RC14668218 COMPANY HEADING

    #lb3b = Label(tk3, image=img3b)
   # lb3b.img3b = img3b #Save reference to image
   # lb3b.place(x=1227,y=1)

    label2 = Label(tk3, text="MEDICAL SOFTWARE FOR ELECTRONIC RECORD KEEEPING", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 

    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)


    #Create Canvas for RC14668218 Irrigation Layout
    canvas5=Canvas(tk3,bg='#FFFFFF',width=1300,height=550)
    canvas5.place(x=35,y=120)
    

    img3=ImageTk.PhotoImage(file="ELELYONC.png")
    lb3 = Label(canvas5, image=img3)
    lb3.img3 = img3 #Save reference to image
    lb3.place(x=20,y=1)

    
    labell1 = Label(canvas5, text="PATIENT  NAME: ", width=17, height=2, font=('Arial bold', 20)) 
    labell1.place(x=200,y=20)
    textBox1a=Text(canvas5, height=4, width=40)
    textBox1a.place(x=520,y=20)
    canvas5.create_rectangle(518, 18, 848, 90,outline= "black", width = 6)

    labell2 = Label(canvas5, text="GENDER: ", width=8, height=2, font=('Arial bold', 20)) 
    labell2.place(x=900,y=20)
    textBox2a=Text(canvas5, height=4, width=20)
    textBox2a.place(x=1060,y=20)
    canvas5.create_rectangle(1058, 18, 1228, 90,outline= "black", width = 6)
    
    labell3 = Label(canvas5, text="LOCATION: ", width=17, height=2, font=('Arial bold', 20)) 
    labell3.place(x=200,y=110)
    textBox3a=Text(canvas5, height=4, width=40)
    textBox3a.place(x=520,y=110)
    canvas5.create_rectangle(518, 108, 848, 180,outline= "black", width = 6)

    labell4 = Label(canvas5, text="AGE: ", width=8, height=2, font=('Arial bold', 20)) 
    labell4.place(x=900,y=110)
    textBox4a=Text(canvas5, height=4, width=20)
    textBox4a.place(x=1060,y=110)
    canvas5.create_rectangle(1058, 108, 1228, 180,outline= "black", width = 6)

    labell5 = Label(canvas5, text="DATE OF VISIT: ", width=17, height=2, font=('Arial bold', 20)) 
    labell5.place(x=200,y=200)
    textBox5a=Text(canvas5, height=4, width=12)
    textBox5a.place(x=520,y=200)
    canvas5.create_rectangle(518, 198, 623, 270,outline= "black", width = 6)

    labell6 = Label(canvas5, text="PRE-EXISTING CONDITION: ", width=22, height=2, font=('Arial bold', 20)) 
    labell6.place(x=640,y=200)
    textBox6a=Text(canvas5, height=4, width=23)
    textBox6a.place(x=1040,y=200)
    canvas5.create_rectangle(1038, 198, 1230, 270,outline= "black", width = 6)


    labell7 = Label(canvas5, text=" DISEASE(S) DIAGNOSED: ", width=20, height=2, font=('Arial bold', 20)) 
    labell7.place(x=5,y=290)
    textBox7a=Text(canvas5, height=4, width=32)
    textBox7a.place(x=360,y=290)
    canvas5.create_rectangle(358, 288, 623, 360,outline= "black", width = 6)

    labell8 = Label(canvas5, text="DRUG(S) PRESCRIBED: ", width=19, height=2, font=('Arial bold', 20)) 
    labell8.place(x=640,y=290)
    textBox8a=Text(canvas5, height=4, width=30)
    textBox8a.place(x=985,y=290)
    canvas5.create_rectangle(983, 288, 1230, 360,outline= "black", width = 6)

    labell9 = Label(canvas5, text="PREVIOUS HEALTH HISTORY: ", width=24, height=2, font=('Arial bold', 20)) 
    labell9.place(x=5,y=400)
    textBox9a=Text(canvas5, height=8, width=64)
    textBox9a.place(x=440,y=390)
    canvas5.create_rectangle(437, 388, 964, 530,outline= "black", width = 6)

    button3 = Button(canvas5, text='ADD DATA',font=('Arial bold', 36), width=10, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=AddPData)
    button3.place(x=980,y=410)

 
   # labell3 = Label(canvas5, text="AGE ", width=5, height=2, font=('Arial bold', 20)) 
   # labell3.place(x=665,y=20)
   # labell4 = Label(canvas5, text="LOCATION ", width=14, height=2, font=('Arial bold', 20)) 
   # labell4.place(x=770,y=20)    
   # labell5 = Label(canvas5, text="DATE OF PHYSICIAN VISIT ", width=30, height=2, font=('Arial bold', 20)) 
   # labell5.place(x=975,y=20)    
   # labell6 = Label(canvas5, text="DISEASE(S) DIAGNOSED ", width=30, height=2, font=('Arial bold', 20)) 
   # labell6.place(x=1165,y=20)    
   # labell7 = Label(canvas5, text="DRUGS PRESCRIBED ", width=40, height=2, font=('Arial bold', 20)) 
   # labell7.place(x=1465,y=20)    
   # labell8 = Label(canvas5, text="PREVIOUS HEALTH HISTORY ", width=60, height=2, font=('Arial bold', 20)) 
  #  labell8.place(x=1765,y=20)    
   # labell9 = Label(canvas5, text="PRE-EXISTING MEDICAL  CONDITION(S)", width=40, height=2, font=('Arial bold', 20)) 
   # labell9.place(x=2065,y=20)






def PSearch():

    global textBox1
    global textBox2
    global foundpatientrow2
    global n
    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("PATIENTDATA2.xlsx")
    # file = open('PATIENTDATA2.xlsx', 'r') 
    # Define variable to read sheet
    dataframe1 = dataframe.active

    #create window
    tk3 = Toplevel()
    tk3.geometry("1200x1200")
    tk3.title('ELECTRONIC RECORD KEEPING BY OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218')
    tk3['background']='#36454F'

    #Insert RC14668218 COMPANY LOGO AND EL ELYON
    # Create a canvas widget
    canvas=Canvas(tk3, width=1400, height=100)
    canvas.pack()

    # Load the images of RC14668218 LOGO AND EL ELYON
   # img2=PhotoImage(file="LOGO3.png")
   # lb2 = Label(tk3, image=img2)
   # lb2.img2 = img2 #Save reference to image
   # lb2.place(x=50,y=50)
   # img3=ImageTk.PhotoImage(file="ELELYONC.png")
   # lb3 = Label(tk3, image=img3)
   # lb3.img3 = img3 #Save reference to image
   # lb3.place(x=1320,y=50)


    #Insert RC14668218 UKFLAG AND UK
    # Load the images of RC14668218 UKFLAG AND UK

    img3b=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND4.png")
    lb3b = Label(tk3, image=img3b)
    lb3b.img3b = img3b #Save reference to image
    lb3b.place(x=1,y=50)
    img3c=ImageTk.PhotoImage(file="MEDSOFTWAREUKN BACKGROUND5.png")
    lb3c = Label(tk3, image=img3c)
    lb3c.img3c = img3c #Save reference to image
    lb3c.place(x=1,y=1)



    #Insert RC14668218 COMPANY HEADING

    #lb3b = Label(tk3, image=img3b)
   # lb3b.img3b = img3b #Save reference to image
   # lb3b.place(x=1227,y=1)

    label2 = Label(tk3, text="MEDICAL SOFTWARE FOR ELECTRONIC RECORD KEEEPING", width=65, height=1, font=('Monotype Corsiva', 18), bg='#ffffff') 

    label2.place(x=260,y=15)
    label3 = Label(tk3, text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Monotype Corsiva', 11), bg='#ffffff') 
    label3.place(x=257,y=60)

    
    #Create Canvas for RC14668218 Irrigation Layout
    canvas5=Canvas(tk3,bg='#FFFFFF',width=1100,height=400)
    canvas5.place(x=50,y=120)
  
    #Create RC14668218 Horizontal and Vertical ScrollBars for Irrigation Layout Section
    hbar = Scrollbar(canvas5, orient = HORIZONTAL)
    hbar.config(command=canvas5.xview)
    hbar.place(x=1, y=440)
    canvas5.configure(xscrollcommand=hbar.set)
    canvas5.config(width=1280,height=455, scrollregion=(10,10,10000,10000))

    vbar = Scrollbar(canvas5, orient = VERTICAL)
    vbar.config(command=canvas5.yview)
    vbar.place(x=1, y=1)
    canvas5.configure(yscrollcommand=vbar.set)
    canvas5.config(width=1280,height=455, scrollregion=(10,10,10000,10000))
       

    
    # irrigation = "You selected the option " + str(radio1.get())
    #Draw RC14668218 IRRIGATION BLOCKS
  
    count1 = 10
    count2 = 20
    count11 = 1
    count12 = 1
    counta = 50
    countb = 140
    countc = 210
    countd = 140
    counte= 50
    countf = 346
    h=0
    i=0
    numarray= [380, 700, 970, 1190, 1580, 2310, 3100, 3940, 5050] 
    numarray2 = 100
    
    img3=ImageTk.PhotoImage(file="ELELYONC.png")
    lb3 = Label(canvas5, image=img3)
    lb3.img3 = img3 #Save reference to image
    lb3.place(x=20,y=1)

    canvas5.create_text(380, 40, text="PATIENT  NAME", width=400, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(700, 40, text="GENDER ", width=300, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(970, 40, text="LOCATION ", width=300, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(1200, 40, text="AGE  ", width=300, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(1580, 40, text="DATE OF PHYSICIAN VISIT  ", width=600, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(2310, 40, text="PRE-EXISTING MEDICAL  CONDITION(S)  ", width=800, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(3060, 40, text="DISEASE(S) DIAGNOSED  ", width=600, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(3960, 40, text="DRUGS PRESCRIBED  ", width=600, font=('Arial bold', 30), fill='#000000')
    canvas5.create_text(5010, 40, text="PREVIOUS HEALTH HISTORY  ", width=600, font=('Arial bold', 30), fill='#000000')
   

    ename = textBox1.get("1.0","end-1c")
    foundpatientrow= 0
    foundpatientrow2= 0
    yfoundpatient = 0
    r=0
    #Iterate the loop to read the cell values MATTHEW RAIN

    #Search  for Patient    
    for row in range(0, dataframe1.max_row):
        if(yfoundpatient == 0):
            for col in dataframe1.iter_cols(1, 1):
                efile = (col[row].value)
                print(col[row].value)
                if(ename == efile):
                    print("FOUND PATIENT!!!!!")
                    yfoundpatient = 0
                    foundpatientrow= row;
                    foundpatientrow2= row;
                    print("PATIENT ROW: ")
                    print(foundpatientrow)
                    foundpatientrow2 = foundpatientrow
                    n=1
                    m=1
                    while (m==1):
                        for col in dataframe1.iter_cols(1, 1):
                            if(foundpatientrow+1<dataframe1.max_row):                               
                                if(col[foundpatientrow+1].value == ename):
                                    n=n+1
                                    r=r+1
                                    foundpatientrow =foundpatientrow+1
                                else:
                                    m=2
                            else:
                                    m=2
                    #List Number of  Records  for Patient                
                    print("NUMBER OF RECORDS: ")
                    print(r)
                    m=1

                    #List Patient Records                 
                    p=0
                    for row in range(foundpatientrow2, dataframe1.max_row):
                         for col in dataframe1.iter_cols(1, dataframe1.max_column):
                             if(row <= (foundpatientrow2 + n - 1)):
                                 print(col[row].value)
                                 canvas5.create_text(numarray[p], numarray2, text=col[row].value, width=600, font=('Arial bold', 20), fill='#000000')
                                 p=p+1
                                 if(p>8):
                                     p=0
                                     numarray2 = numarray2 + 50

                    if(row==dataframe1.max_row):
                        yfoundpatient = 1
                        
    button4 = Button(tk3, text='ADD NEW DATA',font=('Arial bold', 36), width=18, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa',command=PAddNew)
    button4.place(x=615,y=590)


                   # print("FOUND PATIENT ROW!!!!!")
                                                       
                  #      if(row>foundpatientrow):
                   #         for row in range(foundpatientrow+1, dataframe1.max_row):
                   #             for col in dataframe1.iter_cols(1,1):
                      #              if(col[row].value == '*'):
                      #                  for col in dataframe1.iter_cols(1, dataframe1.max_column):
                      #                      print(col[row].value)
                   #         yfoundpatient ==0                
                   #                 #row=dataframe1.max_row+1
                   # break
                   # break
                
                                    
     
  
   # labell1 = Label(canvas5, text="PATIENT  NAME ", width=17, height=2, font=('Arial bold', 20)) 
   # labell1.place(x=200,y=20)
   # labell2 = Label(canvas5, text="GENDER ", width=8, height=2, font=('Arial bold', 20)) 
   # labell2.place(x=510,y=20)
   # labell3 = Label(canvas5, text="AGE ", width=5, height=2, font=('Arial bold', 20)) 
   # labell3.place(x=665,y=20)
   # labell4 = Label(canvas5, text="LOCATION ", width=14, height=2, font=('Arial bold', 20)) 
   # labell4.place(x=770,y=20)    
   # labell5 = Label(canvas5, text="DATE OF PHYSICIAN VISIT ", width=30, height=2, font=('Arial bold', 20)) 
   # labell5.place(x=975,y=20)    
   # labell6 = Label(canvas5, text="DISEASE(S) DIAGNOSED ", width=30, height=2, font=('Arial bold', 20)) 
   # labell6.place(x=1165,y=20)    
   # labell7 = Label(canvas5, text="DRUGS PRESCRIBED ", width=40, height=2, font=('Arial bold', 20)) 
   # labell7.place(x=1465,y=20)    
   # labell8 = Label(canvas5, text="PREVIOUS HEALTH HISTORY ", width=60, height=2, font=('Arial bold', 20)) 
   # labell8.place(x=1765,y=20)    
   # labell9 = Label(canvas5, text="PRE-EXISTING MEDICAL  CONDITION(S)", width=40, height=2, font=('Arial bold', 20)) 
   # labell9.place(x=2065,y=20)
    
   





      
#Insert RC14668218 COMPANY LOGO AND EL ELYON
# Create a canvas widget
canvas=Canvas(mainwindow, width=1400, height=100)
canvas.pack()

# Load the images of RC14668218 LOGO AND EL ELYON
img=ImageTk.PhotoImage(file="COMPANY LOGO A.png")
img2=ImageTk.PhotoImage(file="ELELYON A.png")
# Add the images to the canvas
canvas.create_image(50, 50, image=img)
canvas.create_image(1320, 50, image=img2)

#Insert RC14668218 UKFLAG AND UK
# Load the images of RC14668218 UKFLAG AND UK
img3=ImageTk.PhotoImage(file="UKFLAG.png")
img4=ImageTk.PhotoImage(file="UKMAP.jpeg")
img5=ImageTk.PhotoImage(file="ADCADSOFTWAREUK BACKGROUND4.png")
img6=ImageTk.PhotoImage(file="ADCADSOFTWAREUK BACKGROUND5.png")
img7=ImageTk.PhotoImage(file="ISHI.png")
img8=ImageTk.PhotoImage(file="ELELYON CARD A.png")

# Add the images to the canvas
canvas.create_image(200, 50, image=img3)
canvas.create_image(1235, 50, image=img4)
canvas.create_image(684, 50, image=img6)
canvas.create_image(42, 50, image=img)
canvas.create_image(102, 50, image=img2)
canvas.create_image(198, 50, image=img8)

#Insert RC14668218 COMPANY HEADING
label2 = tk.Label(text="AD CAD SOFTWARE", width=20, height=1, font=('Bodoni MT Black', 28), bg='#808080') 
label2.place(x=560,y=10)
#Bodoni MT Black 
label3 = tk.Label(text="OLAWUYI RACETT NIGERIA LTD., WELLINGTON SQUARE, OXFORD, OX1 2JD, LONDON, UNITED KINGDOM RC14668218", width=98, height=1, font=('Bodoni MT', 11), bg='#808080') 
label3.place(x=357,y=60)



#CREATE SECOND CANVASS
canvas2=Canvas(mainwindow, width=2500, height=700)
canvas2.place(x=1,y=100)
#ADD CANVASS BACKGROND
canvas2.create_image(682, 300, image=img5)

#ADD COMPANY LOGO
#canvas2.create_image(360, 140, image=img)

#ADD ISHI
#canvas3=Canvas(mainwindow, width=240, height=240)
#canvas3.place(x=1,y=1)
#canvas3.create_image(115, 115, image=img)

#create Label and Text Box for PATIENT NAME SEARCH
label4 = tk.Label(text="Select Image for AD CAD Detection :- ", width=30, height=2, font=('Bodoni MT Black', 32), bg='#C0C0C0') 
label4.place(x=20,y=120)


button1 = Button(mainwindow, text='Browse',font=('Arial bold', 32), width=15, height=1, bg='#36454F', fg='#ffffff', bd=6, activebackground='#C0C0C0', activeforeground='#ffffff', command=PSearch)
button1.place(x=900,y=125)







mainwindow.mainloop()


#button1 = Button(mainwindow, text='PATIENT SEARCH',font=('Arial bold', 17), width=16, height=1, bg='#000080', fg='#ffffff', activebackground='#0052cc', activeforeground='#aaffaa', command=power_distribution)
